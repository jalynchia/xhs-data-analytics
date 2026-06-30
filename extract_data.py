#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import argparse
import openpyxl
from typing import Any

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract and compute XHS video metrics.")
    parser.add_argument("--play-file", required=True, help="Path to play data xlsx")
    parser.add_argument("--interact-file", required=True, help="Path to interact data xlsx")
    parser.add_argument("--trend-file", required=True, help="Path to trend JSON")
    parser.add_argument("--visual-json", required=True, help="Path to visual extraction JSON")
    parser.add_argument("--srt-file", required=False, default=None, help="Path to srt subtitle file (optional)")
    parser.add_argument("--out-json", required=True, help="Path to output JSON data")
    return parser.parse_args()

def clean_percentage(val: Any) -> float:
    if val is None:
        return 0.0
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    if val_str.endswith('%'):
        val_str = val_str[:-1]
    try:
        return float(val_str) / 100.0
    except ValueError:
        return 0.0

def clean_seconds(val: Any) -> float:
    if val is None:
        return 0.0
    val_str = str(val).strip()
    if not val_str:
        return 0.0
    if val_str.endswith('s'):
        val_str = val_str[:-1]
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def clean_int(val: Any) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).strip()))
    except ValueError:
        return 0

def extract_summary_sheet(ws: Any) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[0] is not None:
            key = str(row[0]).strip()
            data[key] = row[1]
    return data

def extract_time_series_sheet(ws: Any, clean_func: Any = None) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and row[0] is not None:
            key = str(row[0]).strip()
            val = row[1]
            if clean_func:
                val = clean_func(val)
            data[key] = val
    return data

def srt_time_to_seconds(time_str: str) -> float:
    # formats: HH:MM:SS,mmm or HH:MM:SS.mmm
    time_str = time_str.replace(',', '.')
    parts = time_str.split(':')
    if len(parts) == 3:
        try:
            h = int(parts[0])
            m = int(parts[1])
            s = float(parts[2])
            return h * 3600 + m * 60 + s
        except ValueError:
            return 0.0
    return 0.0

def parse_srt(srt_path: str | None) -> list[dict[str, Any]]:
    subtitles: list[dict[str, Any]] = []
    if not srt_path or not os.path.exists(srt_path):
        return subtitles
    try:
        with open(srt_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"Warning: Failed to read SRT file: {e}", file=sys.stderr)
        return subtitles

    blocks = content.strip().replace('\r\n', '\n').split('\n\n')
    for block in blocks:
        lines = block.split('\n')
        if len(lines) >= 3:
            time_line = lines[1]
            if '-->' in time_line:
                times = time_line.split('-->')
                start_str = times[0].strip()
                end_str = times[1].strip()
                
                start_sec = srt_time_to_seconds(start_str)
                end_sec = srt_time_to_seconds(end_str)
                
                text = " ".join(lines[2:])
                subtitles.append({
                    "start": start_sec,
                    "end": end_sec,
                    "text": text
                })
    return subtitles

def main() -> None:
    args = parse_args()

    # Create output directory if it does not exist
    out_dir = os.path.dirname(args.out_json)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    # 1. Load Workbooks
    wb_play = openpyxl.load_workbook(args.play_file, data_only=True)
    wb_interact = openpyxl.load_workbook(args.interact_file, data_only=True)
    
    # Verify sheet existence
    if "基础数据总览" not in wb_play.sheetnames:
        print(f"❌ 错误：在播放数据文件 {args.play_file} 中找不到名为『基础数据总览』的标签页。\n请检查文件是否为小红书导出的原始格式。", file=sys.stderr)
        sys.exit(1)
    if "互动数据总览" not in wb_interact.sheetnames:
        print(f"❌ 错误：在互动数据文件 {args.interact_file} 中找不到名为『互动数据总览』的标签页。\n请检查文件是否为小红书导出的原始格式。", file=sys.stderr)
        sys.exit(1)

    # 2. Extract play summary
    ws_play_summary = wb_play["基础数据总览"]
    play_summary_raw = extract_summary_sheet(ws_play_summary)
    
    # 3. Extract interact summary
    ws_interact_summary = wb_interact["互动数据总览"]
    interact_summary_raw = extract_summary_sheet(ws_interact_summary)
    
    # Extract title
    import re
    base_name = os.path.basename(args.play_file)
    video_title = re.sub(r'[-_](播放|观看|互动)数据明细表?(\.xlsx)?$', '', base_name, flags=re.IGNORECASE)
    
    # 4. Extract trend JSON
    with open(args.trend_file, 'r', encoding='utf-8') as f:
        trend_raw = json.load(f)
    
    trend_list = sorted(trend_raw["data"].get("trend_list", []), key=lambda x: x["date"])
    similar_trend_list = sorted(trend_raw["data"].get("similar_trend_list", []), key=lambda x: x["date"])
    
    video_duration = 0
    if trend_list:
        video_duration = max(x["date"] for x in trend_list)
        
    # 5. Extract visual JSON
    with open(args.visual_json, 'r', encoding='utf-8') as f:
        visual_data = json.load(f)

    # 6. Parse SRT subtitle if present
    subtitles = parse_srt(args.srt_file)

    # 7. Resolve summary fields & apply naming correction
    total_impressions = clean_int(play_summary_raw.get("曝光数"))
    total_views = clean_int(play_summary_raw.get("观看数"))
    
    ctr_pct = clean_percentage(play_summary_raw.get("封面点击率(%)"))
    avg_watch_sec = clean_seconds(play_summary_raw.get("平均观看时长(s)"))
    completion_pct = clean_percentage(play_summary_raw.get("完播率(%)"))
    exit2s_pct = clean_percentage(play_summary_raw.get("2秒退出率(%)"))
    new_followers = clean_int(play_summary_raw.get("涨粉数"))
    
    fan_impression_pct = clean_percentage(play_summary_raw.get("曝光数粉丝占比(%)"))
    fan_view_pct = clean_percentage(play_summary_raw.get("观看数粉丝占比(%)"))
    
    fan_ctr_pct = clean_percentage(play_summary_raw.get("粉丝-封面点击率(%)") or play_summary_raw.get("粉丝 - 封面点击率(%)") or play_summary_raw.get("封面点击率粉丝占比(%)"))
    fan_avg_watch_sec = clean_seconds(play_summary_raw.get("粉丝-平均观看时长粉丝(%)") or play_summary_raw.get("粉丝 - 平均观看时长 (s)") or play_summary_raw.get("平均观看时长粉丝占比(%)"))
    fan_completion_pct = clean_percentage(play_summary_raw.get("粉丝-完播率(%)") or play_summary_raw.get("粉丝 - 完播率 (%)") or play_summary_raw.get("完播率粉丝占比(%)"))
    fan_exit2s_pct = clean_percentage(play_summary_raw.get("粉丝-2秒退出率(%)") or play_summary_raw.get("粉丝 - 2秒退出率 (%)") or play_summary_raw.get("2秒退出率粉丝占比(%)"))

    # Interact details
    total_likes = clean_int(interact_summary_raw.get("点赞数"))
    total_collects = clean_int(interact_summary_raw.get("收藏数"))
    total_comments = clean_int(interact_summary_raw.get("评论数"))
    total_shares = clean_int(interact_summary_raw.get("分享数"))
    
    # Derived Metrics
    interact_rate = (total_likes + total_collects + total_comments + total_shares) / total_views if total_views > 0 else 0.0
    average_playtime_pct = avg_watch_sec / video_duration if video_duration > 0 else 0.0
    follower_conversion_rate = new_followers / total_views if total_views > 0 else 0.0

    # 8. Match time series data (Day-level)
    dates = []
    play_sheets_day = ["曝光数（天）", "观看数（天）", "封面点击率（天）", "平均观看时长（天）", "完播率（天）", "2s退出率（天）", "涨粉数（天）"]
    day_data_dict = {}
    for sname in play_sheets_day:
        if sname in wb_play.sheetnames:
            day_data_dict[sname] = extract_time_series_sheet(wb_play[sname])
            dates.extend(day_data_dict[sname].keys())
            
    interact_sheets_day = ["点赞（天）", "评论数（天）", "收藏数（天）", "分享数（天）"]
    for sname in interact_sheets_day:
        if sname in wb_interact.sheetnames:
            day_data_dict[sname] = extract_time_series_sheet(wb_interact[sname])
            dates.extend(day_data_dict[sname].keys())
            
    dates = sorted(list(set(dates)))
    time_series_day = []
    for d in dates:
        time_series_day.append({
            "date": d,
            "impressions": clean_int(day_data_dict.get("曝光数（天）", {}).get(d, 0)),
            "views": clean_int(day_data_dict.get("观看数（天）", {}).get(d, 0)),
            "ctr": clean_percentage(day_data_dict.get("封面点击率（天）", {}).get(d, 0.0)),
            "avg_watch_seconds": clean_seconds(day_data_dict.get("平均观看时长（天）", {}).get(d, 0.0)),
            "completion_rate": clean_percentage(day_data_dict.get("完播率（天）", {}).get(d, 0.0)),
            "exit2s_rate": clean_percentage(day_data_dict.get("2s退出率（天）", {}).get(d, 0.0)),
            "new_followers": clean_int(day_data_dict.get("涨粉数（天）", {}).get(d, 0)),
            "likes": clean_int(day_data_dict.get("点赞（天）", {}).get(d, 0)),
            "comments": clean_int(day_data_dict.get("评论数（天）", {}).get(d, 0)),
            "collects": clean_int(day_data_dict.get("收藏数（天）", {}).get(d, 0)),
            "shares": clean_int(day_data_dict.get("分享数（天）", {}).get(d, 0))
        })
        
    time_period = "-"
    if dates:
        time_period = f"{dates[0]} 至 {dates[-1]}"

    # 9. Match time series data (Hour-level)
    hours = []
    play_sheets_hour = ["平均观看时长（小时）", "完播率（小时）", "涨粉数（小时）", "2s退出率（小时）"]
    hour_data_dict = {}
    for sname in play_sheets_hour:
        if sname in wb_play.sheetnames:
            hour_data_dict[sname] = extract_time_series_sheet(wb_play[sname])
            hours.extend(hour_data_dict[sname].keys())
            
    interact_sheets_hour = ["点赞（小时）", "评论数（小时）", "收藏数（小时）", "分享数（小时）"]
    for sname in interact_sheets_hour:
        if sname in wb_interact.sheetnames:
            hour_data_dict[sname] = extract_time_series_sheet(wb_interact[sname])
            hours.extend(hour_data_dict[sname].keys())
            
    hours = sorted(list(set(hours)))
    time_series_hour = []
    for h in hours:
        time_series_hour.append({
            "hour": h,
            "avg_watch_seconds": clean_seconds(hour_data_dict.get("平均观看时长（小时）", {}).get(h, 0.0)),
            "completion_rate": clean_percentage(hour_data_dict.get("完播率（小时）", {}).get(h, 0.0)),
            "new_followers": clean_int(hour_data_dict.get("涨粉数（小时）", {}).get(h, 0)),
            "exit2s_rate": clean_percentage(hour_data_dict.get("2s退出率（小时）", {}).get(h, 0.0)),
            "likes": clean_int(hour_data_dict.get("点赞（小时）", {}).get(h, 0)),
            "comments": clean_int(hour_data_dict.get("评论数（小时）", {}).get(h, 0)),
            "collects": clean_int(hour_data_dict.get("收藏数（小时）", {}).get(h, 0)),
            "shares": clean_int(hour_data_dict.get("分享数（小时）", {}).get(h, 0))
        })

    # 10. Format retention seconds for output
    time_series_retention = []
    trend_map = {x["date"]: x["count"] / 100.0 for x in trend_list}
    similar_map = {x["date"]: x["count"] / 100.0 for x in similar_trend_list}
    all_seconds = sorted(list(set(list(trend_map.keys()) + list(similar_map.keys()))))
    for s in all_seconds:
        time_series_retention.append({
            "second": s,
            "retention": trend_map.get(s, 0.0),
            "similar_retention": similar_map.get(s, 0.0)
        })

    # 11. Prepare fan vs overall comparison
    fan_comparison = [
        {
            "metric_name": "曝光占比",
            "global_val": 1.0 - fan_impression_pct,
            "fan_val": fan_impression_pct,
            "diff": None,
            "unit": "%"
        },
        {
            "metric_name": "观看占比",
            "global_val": 1.0 - fan_view_pct,
            "fan_val": fan_view_pct,
            "diff": None,
            "unit": "%"
        },
        {
            "metric_name": "封面点击率",
            "global_val": ctr_pct,
            "fan_val": fan_ctr_pct,
            "diff": fan_ctr_pct - ctr_pct,
            "unit": "%"
        },
        {
            "metric_name": "平均播放时长",
            "global_val": avg_watch_sec,
            "fan_val": fan_avg_watch_sec,
            "diff": fan_avg_watch_sec - avg_watch_sec,
            "unit": "s"
        },
        {
            "metric_name": "完播率",
            "global_val": completion_pct,
            "fan_val": fan_completion_pct,
            "diff": fan_completion_pct - completion_pct,
            "unit": "%"
        },
        {
            "metric_name": "2秒退出率",
            "global_val": exit2s_pct,
            "fan_val": fan_exit2s_pct,
            "diff": fan_exit2s_pct - exit2s_pct,
            "unit": "%"
        }
    ]

    # 12. Assembly Output Structure
    output_data = {
        "title": video_title,
        "time_period": time_period,
        "video_duration_seconds": video_duration,
        "summary_kpi": {
            "impressions": total_impressions,
            "views": total_views,
            "ctr": ctr_pct,
            "average_playtime_pct": average_playtime_pct,
            "interact_rate": interact_rate,
            "follower_conversion_rate": follower_conversion_rate
        },
        "interact_detail": {
            "likes": total_likes,
            "collects": total_collects,
            "comments": total_comments,
            "shares": total_shares
        },
        "traffic_sources": visual_data.get("traffic_sources", {}),
        "demographics": visual_data.get("demographics", {}),
        "fan_comparison": fan_comparison,
        "time_series_day": time_series_day,
        "time_series_hour": time_series_hour,
        "time_series_retention": time_series_retention,
        "subtitles": subtitles
    }

    # Write output
    with open(args.out_json, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"Data extraction, srt parsing, and metric computation finished. JSON exported to {args.out_json}")

if __name__ == "__main__":
    main()
