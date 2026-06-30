#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import json
import argparse
import openpyxl
from openpyxl.styles import Font
from typing import Any

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate raw data summary Excel file.")
    parser.add_argument("--in-json", required=True, help="Path to input computed JSON data")
    parser.add_argument("--out-xlsx", required=True, help="Path to save summary Excel file")
    return parser.parse_args()

def write_day_sheet(wb: Any, day_data: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("天维度数据汇总")
    
    # Headers
    headers = [
        "日期", "曝光数", "观看数", "封面点击率", 
        "人均观看时长", "完播率", "2s退出率", "涨粉数", 
        "点赞数", "评论数", "收藏数", "笔记分享数"
    ]
    ws.append(headers)
    
    # Formatting headers
    bold_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
    
    # Rows
    for row in day_data:
        ws.append([
            row["date"],
            row["impressions"],
            row["views"],
            row["ctr"],
            row["avg_watch_seconds"],
            row["completion_rate"],
            row["exit2s_rate"],
            row["new_followers"],
            row["likes"],
            row["comments"],
            row["collects"],
            row["shares"]
        ])
        
    # Apply standard number formatting
    # Row 2 to max_row
    for row_idx in range(2, len(day_data) + 2):
        # Date (Col 1) - General
        
        # Impressions (Col 2), Views (Col 3), New Followers (Col 8), Likes (Col 9), Comments (Col 10), Collects (Col 11), Shares (Col 12) - #,##0
        for col in [2, 3, 8, 9, 10, 11, 12]:
            ws.cell(row=row_idx, column=col).number_format = '#,##0'
            
        # CTR (Col 4), Completion Rate (Col 6), Exit Rate (Col 7) - 0.0%
        for col in [4, 6, 7]:
            ws.cell(row=row_idx, column=col).number_format = '0.0%'
            
        # Avg Watch Seconds (Col 5) - 0.0"s" or General
        ws.cell(row=row_idx, column=5).number_format = '0.0'

def write_hour_sheet(wb: Any, hour_data: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("小时维度数据汇总")
    
    headers = [
        "时刻", "人均观看时长(秒)", "完播率(小数)", "涨粉数", 
        "点赞数", "评论数", "收藏数", "笔记分享数"
    ]
    ws.append(headers)
    
    bold_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        
    for row in hour_data:
        ws.append([
            row["hour"],
            row["avg_watch_seconds"],
            row["completion_rate"],
            row["new_followers"],
            row["likes"],
            row["comments"],
            row["collects"],
            row["shares"]
        ])
        
    for row_idx in range(2, len(hour_data) + 2):
        ws.cell(row=row_idx, column=2).number_format = '0.0'
        ws.cell(row=row_idx, column=3).number_format = '0.0%'
        
        for col in [4, 5, 6, 7, 8]:
            ws.cell(row=row_idx, column=col).number_format = '#,##0'

def write_retention_sheet(wb: Any, retention_data: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("观看趋势（秒）")
    
    headers = ["播放进度（秒）", "本视频", "同类视频"]
    ws.append(headers)
    
    bold_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold_font
        
    for row in retention_data:
        ws.append([
            row["second"],
            row["retention"],
            row["similar_retention"]
        ])
        
    for row_idx in range(2, len(retention_data) + 2):
        ws.cell(row=row_idx, column=1).number_format = '#,##0'
        ws.cell(row=row_idx, column=2).number_format = '0.0%'
        ws.cell(row=row_idx, column=3).number_format = '0.0%'

def main() -> None:
    args = parse_args()
    
    with open(args.in_json, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Write Sheets
    write_day_sheet(wb, data.get("time_series_day", []))
    write_hour_sheet(wb, data.get("time_series_hour", []))
    write_retention_sheet(wb, data.get("time_series_retention", []))
    
    # Create output directory if it does not exist
    out_dir = os.path.dirname(args.out_xlsx)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
        
    wb.save(args.out_xlsx)
    print(f"Summary Excel sheet successfully written to {args.out_xlsx}")

if __name__ == "__main__":
    main()
